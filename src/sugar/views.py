# -*- coding: utf-8 -*-
#
from datetime import datetime, timedelta
from django.conf import settings
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404 as getObj
from django.utils import timezone
from django.utils.translation import ugettext
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from webframe.functions import getDateTime, getDate, FMT_DATE, FMT_DATETIME, getEndOfDay
from .models import Record

import logging
logger=logging.getLogger('sugar.views')

def _getUser(req, username=None):
   if username:
      if req.user.is_superuser or req.user.username==username:
         return getObj(get_user_model(), username=username)
   else:
      return req.user
   raise PermissionDenied()

def index(req):
   if req.user.is_authenticated():
      return redirect('dashboard', username=req.user.username)
   return render(req, 'webframe/empty.html')

@login_required
def dashboard(req, username=None):
   user=_getUser(req, username)

   if req.method=='POST':
      with transaction.atomic():
         r=Record()
         r.owner=getObj(get_user_model(), username=username)
         r.date=getDateTime(req.POST.get('date'))
         r.sugar=req.POST.get('sugar', '0')
         r.sugar=float(r.sugar) if r.sugar else 0
         r.pulse=req.POST.get('pulse', '0')
         r.pulse=int(r.pulse) if r.pulse else 0
         r.sys=req.POST.get('sys', '0')
         r.sys=int(r.sys) if r.sys else 0
         r.dia=req.POST.get('dia', '0')
         r.dia=int(r.dia) if r.dia else 0
         r.save()
         return redirect('reports-user', username=username) if username else redirect('reports')

   return render(req, 'sugar/dashboard.html', {})

@login_required
def reports(req, username=None):
   user=_getUser(req, username)

   params=dict()
   params['to']=getDate(req.GET.get('to', None), timezone.now())
   params['to']=getEndOfDay(params['to']) #Due to the system should include the selected date instead
   params['from']=getDate(req.GET.get('from', None), params['to']-timedelta(days=30))
   params['target']=Record.objects.filter(owner=user, date__range=(params['from'], params['to'])).order_by('date')

   return render(req, 'sugar/reports.html', params)

@login_required
def downloads(req, username=None):
   user=_getUser(req, username)

   params=dict()
   params['to']=getDate(req.GET.get('to', None), datetime.now())
   params['from']=getDate(req.GET.get('from', None), params['to']-timedelta(days=30))
   params['target']=Record.objects.filter(owner=user, date__range=(params['from'], params['to'])).order_by('date')
   logger.debug(params['target'])

   filename=ugettext('From %(from)s to %(to)s'%params)

   wb=Workbook()
   ws=wb.active
   ws.merge_cells('A1:G1')
   ws['A1']=filename
   ws['A2']=ugettext('Record.owner')
   ws['B2']=user.get_full_name() if user.get_full_name() else user.username
   ws['A3']=ugettext('from')
   ws['B3']=params['from'].strftime(FMT_DATE)
   ws['A4']=ugettext('to')
   ws['B4']=params['to'].strftime(FMT_DATE)

   ws.cell(row=5, column=3, value=ugettext('Record.date'))
   ws.cell(row=5, column=4, value=ugettext('Record.sugar'))
   ws.cell(row=5, column=5, value=ugettext('Record.pulse'))
   ws.cell(row=5, column=6, value=ugettext('Record.sys'))
   ws.cell(row=5, column=7, value=ugettext('Record.dia'))
   row=6
   for r in params['target']:
      ws.cell(row=row, column=3, value=timezone.localtime(r.date).strftime(FMT_DATETIME))
      ws.cell(row=row, column=4, value=r.sugar)
      ws.cell(row=row, column=5, value=r.pulse)
      ws.cell(row=row, column=6, value=r.sys)
      ws.cell(row=row, column=7, value=r.dia)
      row+=1
   rst=HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
   rst['Content-Disposition'] = 'attachment; filename=\"%s.xlsx\"'%filename
   return rst
